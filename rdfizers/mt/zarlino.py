import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser()
parser.add_argument("--in_xlsx")
args = parser.parse_args()


wb = load_workbook(filename=args.in_xlsx)
ws = wb[wb.sheetnames[0]]

rows = ws.iter_rows(min_row=1, max_row=1)
first_row = next(rows)
headings = [c.value for c in first_row]

data = []

for row in list(ws.rows)[1:]:
    values = [cell.value for cell in row]
    dic = {}
    for i in range(len(headings)):
        if values[i]:
            dic[headings[i]] = values[i]
    data.append(dic)
    if dic:
        print(dic)

wb.close()
