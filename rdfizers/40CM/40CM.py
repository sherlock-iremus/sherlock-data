# TODO SCRIPT A REMANIER

import argparse
from openpyxl import load_workbook
from rdflib import Graph, Namespace, DCTERMS, RDF, RDFS, SKOS, URIRef as u, XSD, Literal as l
import sys
from sherlockcachemanagement import Cache
import requests
import glob
import ntpath
import sys, os
sys.path.append(os.path.abspath(os.path.join('rdfizers/', '')))
# print(sys.path)
from helpers_rdf import *
from helpers_python import *

# Arguments
parser = argparse.ArgumentParser()
parser.add_argument("--collection_id")
parser.add_argument("--excel_coll")
parser.add_argument("--dossier_coll")
parser.add_argument("--excel_index")
parser.add_argument("--output_ttl")
parser.add_argument("--cache_images")
parser.add_argument("--cache_corpus")
parser.add_argument("--cache_personnes")
parser.add_argument("--cache_lieux")
parser.add_argument("--cache_vocab_estampes")
args = parser.parse_args()

init_cache(args)

# Fichiers Excel
# Index des collections
wb_index = load_workbook(args.excel_index)
index = wb_index.active
# Images de la collection
if args.excel_coll:
	wb_img = load_workbook(args.excel_coll)
	img = wb_img.active

#####################################################################
# LA COLLECTION
#####################################################################

collection_row = None

for row in index:
	if row[0].value == args.collection_id:
		collection_row = row
		break

#####################################################################
# La publication
#####################################################################

	# Work
	livre_F1 = she(cache_images.get_uuid(["collection", "livre", "F1"], True))
	t(livre_F1, a, lrm("F1_Work"))
	livre_E35 = she(cache_images.get_uuid(["collection", "livre", "E41"], True))
	t(livre_F1, crm("P102_has_title"), livre_E35)
	t(livre_E35, a, crm("E35_Title"))
	t(livre_E35, RDFS.label, l(collection_row[1].value))

	# Work Conception
	livre_F27 = she(cache_images.get_uuid(["collection", "livre", "F27", "uuid"], True))
	t(livre_F27, a, lrm("F27_Work_Conception"))
	t(livre_F27, lrm("R16_initiated"), livre_F1)
	t(livre_F27, crm("P14_carried_out_by"), l(collection_row[8].value))
	if collection_row[9].value != None:
		livre_F27_E52 = she(cache_images.get_uuid(["collection", "livre", "F27", "E52"], True))
		t(livre_F27, crm("P4_has_time-span"), livre_F27_E52)
		t(livre_F27_E52, crm("P80_end_is_qualified_by"), l(collection_row[9].value))

	# Expression
	livre_F2 = she(cache_images.get_uuid(["collection", "livre", "F2"], True))
	t(livre_F1, lrm("R3_is_realised_in"), livre_F2)
	t(livre_F2, a, lrm("F2_Expression"))
	# Expression Creation
	livre_F28 = she(cache_images.get_uuid(["collection", "livre", "F28", "uuid"], True))
	t(livre_F28, a, lrm("F28_Expression_Creation"))
	t(livre_F28, lrm("R17_created"), livre_F2)
	t(livre_F28, crm("P14_carried_out_by"), l(collection_row[8].value))
	if collection_row[10].value != None:
		livre_F28_E52 = she(cache_images.get_uuid(["collection", "livre", "F28", "E52"], True))
		t(livre_F28, crm("P4_has_time-span"), livre_F28_E52)
		t(livre_F28_E52, crm("P80_end_is_qualified_by"), l(collection_row[10].value))

	# Manifestation
	livre_F3 = she(cache_images.get_uuid(["collection", "livre", "F3"], True))
	t(livre_F3, a, lrm("F3_Manifestation"))
	t(livre_F3, lrm("R4_embodies"), livre_F2)
	## Manifestation Creation
	livre_F30 = she(cache_images.get_uuid(["collection", "livre", "F30", "uuid"], True))
	t(livre_F30, a, lrm("F30_Manifestation_Creation"))
	t(livre_F30, lrm("R24_created"), livre_F3)
	t(livre_F30, crm("P92_brought_into_existence"), livre_F2)
	if collection_row[11].value != None:
		livre_F30_E52 = she(cache_images.get_uuid(["collection", "livre", "F30", "E52"], True))
		t(livre_F30, crm("P4_has_time-span"), livre_F30_E52)
		t(livre_F30_E52, crm("P80_end_is_qualified_by"), l(collection_row[11].value))
	# Item
	livre_F5 = she(cache_images.get_uuid(["collection", "livre", "F5"], True))
	t(livre_F5, a, lrm("F5_Item"))
	t(livre_F5, lrm("R7_is_materialization_of"), livre_F3)

	#####################################################################
	# Les pages de la publication
	#####################################################################

	img_row = None

	for row in img:
		if row[1].value == args.collection_id:
			img_row = row
			id = img_row[0].value

			# La page comme support physique
			page_E18 = she(cache_images.get_uuid(["collection", "livre", "pages", id, "E18"], True))
			t(page_E18, a, crm("E18_Physical_Object"))
			t(livre_F5, crm("P46_is_composed_of"), page_E18)

			# La page comme support sémiotique
			page_E90 = she(cache_images.get_uuid(["collection", "livre", "pages", id, "E90"], True))
			t(page_E90, a, crm("E90_Symbolic_Object"))
			t(livre_F2, lrm("R15_has_fragment"), page_E90)
			t(page_E18, crm("P128_carries"), page_E90)

			# Identifiant
			page_id = she(cache_images.get_uuid(["collection", "livre", "pages", id, "E42", "id"], True))
			t(page_E90, crm("P1_is_identified_by"), page_id)
			t(page_id, a, crm("E42_Identifier"))
			t(page_id, RDFS.label, l(id))

			# Numéro de la page
			page_no = she(cache_images.get_uuid(["collection", "livre", "pages", id, "E42", "numéro"], True))
			t(page_no, crm("P2_has_type"), she("466bb717-b90f-4104-8f4e-5a13fdde3bc3"))
			t(page_E90, crm("P1_is_identified_by"), page_no)
			t(page_no, a, crm("E42_Identifier"))
			t(page_no, RDFS.label, l(f"Page {img_row[3].value}"))

			# Numérisation de la page
			page_D2 = she(cache_images.get_uuid(["collection", "livre", "pages", "D2"], True))
			t(page_D2, a, crmdig("D2_Digitization_Process"))
			t(page_D2, crmdig("L1_digitized"), page_E18)
			page_D1 = she(cache_images.get_uuid(["collection", "livre", "pages", id, "D1"], True))
			t((page_D1), a, crmdig("D1_Digital_Object"))
			t(page_D2, crmdig("L11_had_output"), page_D1)
			t(page_D1, crm("130_shows_features_of"), page_E90)
			t(collection, crm("P106_is_composed_of"), page_D1)

			# TODO Transcription de la page

####################################################################################
# Ecriture du graphe
####################################################################################

serialization = output_graph.serialize(format="turtle", base="http://data-iremus.huma-num.fr/id/")
with open(args.output_ttl, "wb") as f:
	f.write(serialization)

cache_images.bye()
