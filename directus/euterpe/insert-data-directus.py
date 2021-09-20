import os, sys
import requests
import yaml
import argparse
from openpyxl import load_workbook
from pprint import pprint
import time

# Helpers
sys.path.append(os.path.abspath(os.path.join('rdfizers/', '')))
# print(sys.path)
from helpers_python import *

# Arguments
parser = argparse.ArgumentParser()
parser.add_argument("--data")
parser.add_argument("--taxonomies")
args = parser.parse_args()

# Secret YAML
file = open(os.path.join(sys.path[0], "secret.yaml"))
secret = yaml.full_load(file)
r = requests.post(secret["url"] + "/auth/login",
                  json={"email": secret["email"], "password": secret["password"]})
access_token = r.json()['data']['access_token']
refresh_token = r.json()['data']['refresh_token']
file.close()


################################################################################################
## FONCTIONS
################################################################################################

# Suppression d'une collection Directus
def delete(collection):
	# Création d'une liste des identifiants des données à supprimer grâce à une requête GET
	r = requests.get(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token)
	print("Récupération des données:", r)
	ids = [item["id"] for item in r.json()["data"]]

	# Suppression des données par parquets de 100
	try:
		for i in range(0, len(ids), 100):
			ids_slice = [ids[j] for j in range(i, i + 100) if j < len(ids)]
			print(i)
			try:
				r = requests.delete(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token, json=ids_slice)
				print("Suppression des données par paquets de 100 :", r)
			except Exception as e:
				print(e)
			n = i

		# Suppression des données restantes (non envoyées car elles n'atteignent pas la centaine de données)
		try:
			for i in range(n, len(ids), 1):
				print(i)
				try:
					r = requests.delete(
						secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token, json=ids[i])
					print("Suppression des données restantes :", r)
				except Exception as e:
					print(e)
		except:
			pass
	except:
		pass


# Création d'une collection Directus à partir du fichier "taxonomies.xlsx"
def send_taxonomy(sheet, collection):
	rows = get_xlsx_sheet_rows_as_dicts(taxonomies[sheet])
	for row in rows:
		if row["name"] != None:
			# Création d'un dictionnaire par ligne de feuille Excel
			dict = {"id": row["uuid"], "nom": row["name"]}

			# Ajout de la correspondance identifiant_euterpe-UUID au dictionnaire "id_uuid"
			id_uuid[str(row["id"])] = row["uuid"]

			# Insertion du dictionnaire dans la collection Directus
			# r = requests.post(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token, json=dict)
			# print(r)


# Création d'une collection Directus à partir du fichier "euterpe_data.xlsx"
def send_data(collection, paquet, range_min, range_max):
	print("Données à insérer:", len(data_to_send))

	# Envoi des données par paquets de 100
	for i in range(0, len(data_to_send), paquet):
		data_slice = [data_to_send[j] for j in range(i, i + paquet) if j < len(data_to_send)]
		print(i)
		try:
			r = requests.post(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token, json=data_slice)
			r.raise_for_status()
		except Exception as e:
			print(e)
			pprint(data_slice)
		print(r)
		print("\n")
		# time.sleep(2)

	# Envoi des données restantes (non envoyées car elles n'atteignent pas la centaine de données)
	for i in range(range_min, range_max):
		print(i)
		try:
			r = requests.post(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token, json=data_to_send[i])
			r.raise_for_status()
		except Exception as e:
			print(e)
			# pprint(data_to_send[i])
		print(r)
		print("\n")
		# time.sleep(2)


# Fonction de recherche de l'UUID d'un concept à partir de son identifiant Euterpe,
# en utilisant le dictionnaire "id_uuid"
def get_uuid_list(column_name, uuid_list):
	row[column_name] = str(row[column_name])
	if "🍄" in row[column_name]:
		ids = row[column_name].split("🍄")
		for id in ids:
			try:
				uuid = id_uuid[id.strip()]
				uuid_list.append(uuid)
			except:
				print(column_name, ":", id, "- id non trouvé")
	else:
		try:
			id = row[column_name]
			uuid = id_uuid[id.strip()]
			uuid_list.append(uuid)
		except:
			print(column_name, ":", row[column_name], "- id non trouvé")


################################################################################################
## TAXONOMIES
################################################################################################

# Lecture du fichier Excel
taxonomies = load_workbook(args.taxonomies)
taxonomies_sheets = taxonomies.sheetnames

# Dictionnaire de correspondance identifiant_euterpe-UUID
id_uuid = {}

# Insertion des données dans Directus
for sheet in taxonomies_sheets:
	if sheet == "spécialité":
		print("SPECIALITES")
		send_taxonomy(sheet, "specialites")
		print("\n" * 2)
	if sheet == "Période":
		print("PERIODES")
		send_taxonomy(sheet, "periodes")
		print("\n" * 2)
	if sheet == "École":
		print("ECOLES")
		send_taxonomy(sheet, "ecoles")
		print("\n" * 2)
	if sheet == "Domaine":
		print("DOMAINES")
		send_taxonomy(sheet, "domaines")
		print("\n" * 2)
	if sheet == "Lieu de conservation":
		print("LIEU DE CONSERVATION")
		send_taxonomy(sheet, "lieux_de_conservation")
		print("\n" * 2)
	if sheet == "Thème":
		print("THEMES")
		send_taxonomy(sheet, "themes")
		print("\n" * 2)
	if sheet == "Instrument de musique":
		print("INSTRUMENTS DE MUSIQUE")
		send_taxonomy(sheet, "instruments_de_musique")
		print("\n" * 2)
	if sheet == "Chant":
		print("CHANTS")
		send_taxonomy(sheet, "chants")
		print("\n" * 2)
	if sheet == "Support":
		print("SUPPORTS")
		send_taxonomy(sheet, "supports")
		print("\n" * 2)
	if sheet == "Type oeuvre":
		print("TYPES D'OEUVRES")
		send_taxonomy(sheet, "types_doeuvres")
		print("\n" * 2)
	if sheet == "Rôles":
		print("ROLES")
		send_taxonomy(sheet, "roles")
		print("\n" * 2)
	if sheet == "Notation musicale":
		print("NOTATION MUSICALE")
		send_taxonomy(sheet, "notation_musicale")
		print("\n" * 2)


################################################################################################
## DATA
################################################################################################

# Lecture du fichier Excel
data = load_workbook(args.data)
data_sheets = data.sheetnames


# 1. AUTEURS (collection "auteurs_oeuvres" dans Directus)
#----------------------------------------------------------

data_to_send = []

rows = get_xlsx_sheet_rows_as_dicts(data["1_auteurs"])

# Suppression de la collection Directus
# delete("auteurs_oeuvres")

for row in rows:

	periodes = []
	specialites = []
	ecoles = []

	# Recherche de l'UUID des "siècles" à partir de leur identifiant euterpe
	if row["siècle"] != None:
		get_uuid_list("siècle", periodes)

	# Recherche de l'UUID des "spécialités" à partir de leur identifiant euterpe
	if row["spécialité"] != None:
		get_uuid_list("spécialité", specialites)

	# Recherche de l'UUID des "écoles" à partir de leur identifiant euterpe
	if row["école"] != None:
		get_uuid_list("école", ecoles)

	# Ajout de la correspondance identifiant_euterpe-UUID des auteurs dans le dictionnaire
	id_uuid[str(row["id"])] = row["uuid"]

	dict = {
		"id": row["uuid"],
		"nom": row["nom"],
		"alias": row["alias"],
		"lieu_de_deces": row["lieu de décès"],
		"periode": [{
			"periodes_id": periode,
			"auteurs_oeuvres_id": row["uuid"],
			"collection": "periode"
		} for periode in periodes],
		"specialite": [{
			"specialites_id": specialite,
			"auteurs_oeuvres_id": row["uuid"],
			"collection": "specialite"
		} for specialite in specialites],
		"ecole": [{
			"ecoles_id": ecole,
			"auteurs_oeuvres_id": row["uuid"],
			"collection": "ecole"
		} for ecole in ecoles],
		"date_de_deces": row["date de décès"],
		"lieu_de_naissance": row["lieu de naissance"],
		"date_de_naissance": row["date de naissance"],
		"commentaire": row["commentaire"],
		"lieu_dactivite": row["lieu d'activité"],
		"date_dactivite": row["date d'activité"]
	}


	# Ajout du dictionnaire dans la liste de données à envoyer
	data_to_send.append(dict)

#Envoi des données dans Directus
# send_data("auteurs_oeuvres", 100, 3300, 3385)


# 2. OEUVRES LYRIQUES
#------------------------

data_to_send = []

rows = get_xlsx_sheet_rows_as_dicts(data["5_oeuvres_lyriques"])

# Suppression de la collection Directus
# delete("oeuvres_lyriques")

for row in rows:

	librettistes = []
	compositeurs = []
	types_oeuvres = []

	# Recherche de l'UUID des "librettistes" à partir de leur identifiant euterpe
	if row["librettiste"] != None:
		get_uuid_list("librettiste", librettistes)

	# Recherche de l'UUID des "compositeurs" à partir de leur identifiant euterpe
	if row["compositeur"] != None:
		get_uuid_list("compositeur", compositeurs)

	# Recherche de l'UUID des "types d'oeuvres" à partir de leur identifiant euterpe
	if row["type_oeuvre"] != None:
		try:
			type_oeuvre_uuid = id_uuid[str(row["type_oeuvre"])]
		except:
			print("type_oeuvre :", row["type_oeuvre"], "non trouvé")

	# Ajout de la correspondance identifiant_euterpe-UUID des oeuvres lyriques dans le dictionnaire
	id_uuid[str(row["id"])] = row["uuid"]

	dict = {
		"id": row["uuid"],
		"titre": row["titre"],
		"librettiste": [{
			"auteurs_oeuvres_id": librettiste,
			"oeuvres_lyriques_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for librettiste in librettistes],
		"compositeur": [{
			"auteurs_oeuvres_id": compositeur,
			"oeuvres_lyriques_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for compositeur in compositeurs],
		"date_oeuvre": row["date_oeuvre"],
		"type_oeuvre": type_oeuvre_uuid,
		"commentaire": row["commentaire"]
	}

	#Ajout du dictionnaire dans la liste de données à envoyer
	data_to_send.append(dict)

# Envoi des données dans Directus
# send_data("oeuvres_lyriques", 100, 100, 123)


# 3. AUTEURS BIBLIOGRAPHIE
#---------------------------

data_to_send = []

rows = get_xlsx_sheet_rows_as_dicts(data["6_auteurs_bibli_id"])

# Suppression de la collection Directus
# delete("auteurs_bibliographie")

for row in rows:

	# Ajout de la correspondance identifiant_euterpe-UUID des oeuvres lyriques dans le dictionnaire
	id_uuid[str(row["id"])] = row["uuid"]

	dict = {
		"id": row["uuid"],
		"nom": row["nom"],
		"prenom": row["prénom"]
	}

	# Ajout du dictionnaire dans la liste de données à envoyer
	data_to_send.append(dict)

# Envoi des données dans Directus
# send_data("auteurs_bibliographie", 100, 400, 436)



# 4. BIBLIOGRAPHIE
#--------------------

data_to_send = []

rows = get_xlsx_sheet_rows_as_dicts(data["3_euterpe_biblio"])

# Suppression de la collection Directus
# delete("bibliographie")

for row in rows:

	auteurs = []

	# Recherche de l'UUID des "auteurs" à partir de leur identifiant euterpe
	if row["auteur_id"] != None:
		get_uuid_list("auteur_id", auteurs)

	# Ajout de la correspondance identifiant_euterpe-UUID des oeuvres lyriques dans le dictionnaire
	id_uuid[str(row["id"])] = row["uuid"]

	dict = {
		"id": row["uuid"],
		"auteur": [{
			"auteurs_bibliographie_id": auteur,
			"bibliographie_id": row["uuid"],
			"collection": "auteurs_bibliographie"
		} for auteur in auteurs],
		"titre": row["titre"],
		"revue_colloque_collection": row["revue_colloque_collection"],
		"parution": row["parution"],
		"editeur": row["editeur"],
		"nbre_n_de_pages": row["nbre_n_de_page"],
		"n_revue_collection": row["n_revue_collection"],
		"lieu_de_publication": row["lieu_d_activite"],
		"commentaire": row["commentaire"]
	}

	# Ajout du dictionnaire dans la liste de données à envoyer
	data_to_send.append(dict)

# Envoi des données dans Directus
send_data("bibliographie", 100, 700, 721)


# 5. OEUVRES
#--------------------

data_to_send = []

rows = get_xlsx_sheet_rows_as_dicts(data["4_euterpe_images"])

# Suppression de la collection Directus
# delete("oeuvres")

for row in rows:

	# Listes qui accueilleront les uuid des concepts indexés
	editeurs = []
	domaines = []
	instruments = []
	inventeurs = []
	themes = []
	lieux_conservation = []
	notations_musicales = []
	graveurs = []
	artistes = []
	chants = []
	attributions = []
	ecoles = []
	anciennes_attributions = []
	ateliers = []
	copies = []
	dapres_list = []
	manieres = []

	# Recherche de l'UUID des concepts indexés à partir de leur identifiant euterpe
	if row["éditeur"] != None:
		get_uuid_list("éditeur", editeurs)

	if row["domaine"] != None:
		get_uuid_list("domaine", domaines)

	if row["instrument de musique"] != None:
		get_uuid_list("instrument de musique", instruments)

	if row["inventeur"] != None:
		get_uuid_list("inventeur", inventeurs)

	if row["thème"] != None:
		get_uuid_list("thème", themes)

	if row["lieu de conservation"] != None:
		get_uuid_list("lieu de conservation", lieux_conservation)

	if row["musique écrite"] != None:
		get_uuid_list("musique écrite", notations_musicales)

	if row["graveur"] != None:
		get_uuid_list("graveur", graveurs)

	if row["artiste"] != None:
		get_uuid_list("artiste", artistes)

	if row["chant"] != None:
		get_uuid_list("chant", chants)

	if row["attribution"] != None:
		get_uuid_list("attribution", attributions)

	if row["école"] != None:
		get_uuid_list("école", ecoles)

	if row["ancienne attribution"] != None:
		get_uuid_list("ancienne attribution", anciennes_attributions)

	if row["atelier"] != None:
		get_uuid_list("atelier", ateliers)

	if row["copie d'après"] != None:
		get_uuid_list("copie d'après", copies)

	if row["d'après"] != None:
		get_uuid_list("d'après", dapres_list)

	if row["manière de"] != None:
		get_uuid_list("manière de", manieres)

	# Ajout de la correspondance identifiant_euterpe-UUID des oeuvres lyriques dans le dictionnaire
	id_uuid[str(row["id"])] = row["uuid"]
	dict = {
		"id": row["uuid"],
		"titre": row["titre"],
		"titre_alternatif": row["titre alternatif"],
		"editeur": [{
			"auteurs_oeuvres_id": editeur,
			"oeuvres_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for editeur in editeurs],
		"reference_iremus": row["référence iremus"].replace("🍄", ",") if row["référence iremus"] != None else row["référence iremus"],
		"domaine": [{
			"domaines_id": domaine,
			"oeuvres_id": row["uuid"],
			"collection": "domaines"
		} for domaine in domaines],
		"num_inventaire": row["n° inventaire"],
		"cote": row["cote"],
		"inscription": row["inscription"],
		"technique": row["technique"],
		"oeuvre_en_rapport": row["œuvre en rapport"],
		"themes": [{
			"themes_id": theme,
			"oeuvres_id": row["uuid"],
			"collection": "themes"
		} for theme in themes],
		"inventeur": [{
			"auteurs_oeuvres_id": inventeur,
			"oeuvres_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for inventeur in inventeurs],
		"lieu_de_conservation": [{
			"lieux_de_conservation_id": lieu_conservation,
			"oeuvres_id": row["uuid"],
			"collection": "lieux_de_conservation"
		} for lieu_conservation in lieux_conservation],
		"date_oeuvre": row["date œuvre"],
		"precision_oeuvre": row["précision œuvre"],
		"precision_instrument": row["précision instrument"],
		"notation_musicale": [{
			"notation_musicale_id": notation_musicale,
			"oeuvres_id": row["uuid"],
			"collection": "notation_musicale"
		} for notation_musicale in notations_musicales],
		"graveur": [{
			"auteurs_oeuvres_id": graveur,
			"oeuvres_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for graveur in graveurs],
		"commentaire": row["commentaire"],
		"bibliographie": row["bibliographie"],
		"reference_agence": row["référence agence"],
		"url": row["url"],
		"titre_url": row["titre de l'url"],

		"hauteur": row["hauteur"],
		"largeur": row["largeur"],
		"diametre": row["diamètre"],
		"chant": [{
			"chants_id": chant,
			"oeuvres_id": row["uuid"],
			"collection": "chants"
		} for chant in chants],
		"artiste": [{
			"auteurs_oeuvres_id": artiste,
			"oeuvres_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for artiste in artistes],
		"ecole": [{
				"auteurs_oeuvres_id": ecole,
				"oeuvres_id": row["uuid"],
				"collection": "auteurs_oeuvres"
			} for ecole in ecoles],
		"attribution": [{
					"auteurs_oeuvres_id": attribution,
					"oeuvres_id": row["uuid"],
					"collection": "auteurs_oeuvres"
				} for attribution in attributions],
		"precision_musique": row["précision musique"],
		"ancienne_attribution": [{
			"auteurs_oeuvres_id": ancienne_attribution,
			"oeuvres_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for ancienne_attribution in anciennes_attributions],
		"atelier": [{
			"auteurs_oeuvres_id": atelier,
			"oeuvres_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for atelier in ateliers],
		"source_litteraire": row["source littéraire"],
		"copie_dapres": [{
			"auteurs_oeuvres_id": copie,
			"oeuvres_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for copie in copies],
		"dapres": [{
			"auteurs_oeuvres_id": dapres,
			"oeuvres_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for dapres in dapres_list],
		"a_la_maniere_de": [{
			"auteurs_oeuvres_id": maniere,
			"oeuvres_id": row["uuid"],
			"collection": "auteurs_oeuvres"
		} for maniere in manieres]
	}

	#TODO images et instruments musique
	# "contient/contenu dans" à écrire à la main? (un seul enregistrement)

	# Ajout du dictionnaire dans la liste de données à envoyer
	data_to_send.append(dict)

# Envoi des données dans Directus
# send_data("oeuvres", 10, 10600, 10692)


# "instruments_de_musique": [{
# 	"instruments_de_musique_id": instrument,
# 	"oeuvres_id": row["uuid"],
# 	"collection": "instruments_de_musique"
# } for instrument in instruments]
