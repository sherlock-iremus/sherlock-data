import os, sys
import requests
import yaml
import argparse
from openpyxl import load_workbook
from pprint import pprint
import time
import uuid
import re
from pprint import pprint
from geopy.geocoders import Nominatim
import json

# Helpers
sys.path.append(os.path.abspath(os.path.join('python_packages/helpers_excel', '')))
from helpers_excel import *

# Arguments
parser = argparse.ArgumentParser()
parser.add_argument("--taxonomy_terms_uuid")
parser.add_argument("--excel_data")
parser.add_argument("--excel_taxonomies")
parser.add_argument("--oeuvres_a_envoyer")
args = parser.parse_args()

# Secret YAML
file = open(os.path.join(sys.path[0], "secret.yaml"))
secret = yaml.full_load(file)
r = requests.post(secret["url"] + "/auth/login",
                  json={"email": secret["email"], "password": secret["password"]})
access_token = r.json()['data']['access_token']
refresh_token = r.json()['data']['refresh_token']
file.close()


################################################################################################
## FONCTIONS
################################################################################################

# Supprimer une collection Directus
def delete(collection):
    print(f"\n- Suppression de la collection {collection}:")
    try:
        # Requête GET listant les identifants des items de la collection Directus
        r = requests.get(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token)
        print("Récupération des données:", r)
        ids = [item["id"] for item in r.json()["data"]]
        print(len(ids), "items trouvés")

        # Suppression des items par paquets de 100
        try:
            print("Suppression des items par paquets de 100")
            for i in range(0, len(ids), 100):
                ids_slice = [ids[j] for j in range(i, i + 100) if j < len(ids)]
                try:
                    r = requests.delete(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token,
                                        json=ids_slice)
                    print("Suppression des données par paquets de 100 :")
                    print(i, r)
                except Exception as e:
                    print(e)
                n = i
        except Exception as e:
            n = 0
            print(e)

        # Suppression du reste des items (échappant aux paquets de 100)
        try:
            print("Suppression des items restants")
            for i in range(n, len(ids), 1):
                try:
                    r = requests.delete(
                        secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token, json=ids[i])
                    print("Suppression des données restantes :")
                    print(i, r)
                except Exception as e:
                    print(e)
        except Exception as e:
            print(e)
    except Exception as e:
        print(e)
        print(r.json())


# Ajout de la correspondance entre l'identifiant Euterpe d'un objet et son UUID dans le dictionnaire "id_uuid"
def add_id_uuid(sheet):
    rows = get_xlsx_sheet_rows_as_dicts(excel_taxonomies[sheet])
    for row in rows:
        if row["name"] != None:
            id_uuid[str(row["id"])] = row["uuid"]


# Création d'une collection à partir d'une taxonomie ("taxonomies.xlsx")
def send_taxonomy(sheet, collection):
    rows = get_xlsx_sheet_rows_as_dicts(excel_taxonomies[sheet])
    for row in rows:
        if row["name"] != None:
            # Un dictionnaire par ligne de feuille Excel
            dict = {"id": row["uuid"], "nom": row["name"]}

            # Récupération des coordonnées géographiques des lieux de conservation
            if sheet == "Lieu de conservation":

                geolocator = Nominatim(user_agent="Iremus")

                try:
                    location = geolocator.geocode(row["name"])
                    latitude = str(location.latitude)
                    longitude = str(location.longitude)
                    dict["coordonnees_geographiques"] = {"coordinates": [longitude, latitude], "type": "Point"}
                except:
                    try:
                        city_split = row["name"].split(",")
                        city = city_split[0]
                        location = geolocator.geocode(city)
                        latitude = str(location.latitude)
                        longitude = str(location.longitude)
                        dict["coordonnees_geographiques"] = {"coordinates": [longitude, latitude], "type": "Point"}
                    except:
                        print("Coordonnées du lieu", row["name"], "non trouvées")

        # Envoi des items dans la collection Directus
        try:
            r = requests.post(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token, json=dict)
            print(r)
            print(dict)
            print("\n")
        except Exception as e:
            print(e)

def send_tree_taxonomy(sheet, collection, themes=True):
    rows = get_xlsx_sheet_rows_as_dicts(excel_taxonomies[sheet])
    dicts_a_envoyer = []
    n = 0

    # Envoi des termes sans parents dans Directus
    print("\nEnvoi des termes dans Directus sans leur parent")
    for row in rows:
        if row["name"] != None:
            id = row["name"].split("- ")[0].strip()

            dict = {"id": row["uuid"], "nom": row["name"]}
            dicts_a_envoyer.append(dict)

    print(len(dicts_a_envoyer), "items à envoyer")
    for d in dicts_a_envoyer[n:]:
        print(n)
        try:
            r = requests.post(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token,
                              json=d)
            print(r.json(), "\n")
            n += 1
        except Exception as e:
            print(e)
    
    # Patch des parents
    n = 0
    # Je récupère les données de Directus
    print("\nRécupération des données dans Directus")
    r = requests.get(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token)
    print(r)

    id_uuid = {}
    for terme in r.json()["data"]:
        id_uuid[terme["nom"].split("-")[0].strip()] = terme["id"]

    print("\nRequêtes PATCH pour ajouter leurs parents aux termes")
    for row in rows[n:]:
        n += 1
        if row["name"] != None:
            id = row["name"].split("- ")[0].strip()

            # S'il s'agit d'identifiants Euterpe, on les ignore
            id_euterpe = re.findall("^[^0-9]*$", id)
            if len(id_euterpe) >= 1:
                continue
            else:
                # Récupération des différentes parties de l'identifiant pour retrouver son parent
                print("")
                print(n, ":", row["name"])
                regex = re.compile(r"([0-9a-zA-Z\.\-]+|\(.*?\))")

                def concat_ancestors(l):
                    for i in range(len(l)):
                        if i >= 1:
                            l[i] = l[i - 1] + l[i]
                    return l

                def make_id_fragment_ancestors(id_fragment):
                    l = []
                    if themes == True:
                        if id_fragment.startswith("(+"):
                            numbers = id_fragment[2:-1]
                            l = ["(+" + s + ")" for s in concat_ancestors(list(numbers))]
                        elif id_fragment[0] == "(":
                            l = ["(...)", id_fragment]
                    else:
                        l = concat_ancestors(list(id_fragment))
                    return l

                def make_ancestor_list(id):
                    matches = re.findall(regex, id)
                    res = []
                    last_ancestor = ""
                    for i in range(len(matches)):
                        ancestors = make_id_fragment_ancestors(matches[i])
                        ancestors = [last_ancestor + s for s in ancestors]
                        last_ancestor = ancestors[-1]

                        res += ancestors
                    
                    # Je vérifie que les ancêtres existent dans Directus et je les ajoute s'ils n'y sont pas
                    # Puis je patche chaque ancêtre de la liste à son parent
                    for i in range(len(res)):
                        terme_courant = res[i]
                        # Ignorer les points de la classification Hornbostel-Sachs
                        if terme_courant.endswith(".") or terme_courant.endswith("-"):
                            continue
                        else:
                            if res[i-1].endswith(".") or res[i-1].endswith("-"):
                                parent = res[i-2]
                                print(terme_courant, parent)
                            else:
                                parent = res[i-1]                        

                        try:
                            uuid_parent = str(id_uuid[parent])
                            print("Le terme parent existe déjà dans la base : Requête PATCH")
                            try:
                                dict = {"parent": uuid_parent}
                                r = requests.patch(
                                    secret["url"] + f'/items/{collection}/' + id_uuid[terme_courant] + '?access_token=' + access_token,
                                    json=dict)
                                print(r)
                            except Exception as e:
                                print(e)
                                print(r.json())
                        
                        except:
                            uuid_parent = str(uuid.uuid4())
                            print("Le terme parent n'existe pas dans la base : envoi dans Directus")
                            try:
                                r = requests.post(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token,
                                                    json={"id": uuid_parent, "nom": parent})
                                print(r.json(), "\n")
                                id_uuid[parent] = uuid_parent

                            except Exception as e:
                                print(e)  

                            # Patch 
                            try:
                                dict = {"parent": uuid_parent}
                                r = requests.patch(
                                    secret["url"] + f'/items/{collection}/' + id_uuid[terme_courant] + '?access_token=' + access_token,
                                    json=dict)
                                print(r)
                            except Exception as e:
                                print(e)
                                print(r.json())                         
                                
                        else:
                            pass
                                    
                make_ancestor_list(id)
                
#Création d'une collection Directus à partir de "euterpe_data.xlsx"
def send_data(collection, paquet, range_min, range_max):
    print("Envoi de", len(donnees_a_envoyer), "items")

    # Envoi des items par paquets de 100
    for i in range(0, len(donnees_a_envoyer), paquet):
        data_slice = [donnees_a_envoyer[j] for j in range(i, i + paquet) if j < len(donnees_a_envoyer)]
        print(i)
        try:
            r = requests.post(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token, json=data_slice)
            r.raise_for_status()
        except Exception as e:
            print(e)
            pprint(data_slice)
            print(r.json())
            print("\n")

    # Envoi des items restants (échappant aux paquets de 100)
    for i in range(range_min, range_max):
        print(i)
        try:
            r = requests.post(secret["url"] + f'/items/{collection}?limit=-1&access_token=' + access_token, json=donnees_a_envoyer[i])
            r.raise_for_status()
        except Exception as e:
            print(e)
            # pprint(donnees_a_envoyer[i])
            print(r.json())
            print("\n")


# Recherche des UUID de plusieurs objets à partir de leur identifiant Euterpe,
# grâce au dictionnaire de correspondance créé précedemment
def get_uuid_list(column_name, uuid_list):
    row[column_name] = str(row[column_name])
    if "🍄" in row[column_name]:
        ids = row[column_name].split("🍄")
        for id in ids:
            try:
                uuid = id_uuid[id.strip()]
                uuid_list.append(uuid)
            except:
                print(column_name, ":", id, "- id non trouvé")

    else:
        try:
            id = row[column_name]
            uuid = id_uuid[id.strip()]
            uuid_list.append(uuid)
        except:
            print(column_name, ":", row[column_name], "- id non trouvé")


################################################################################################
## TAXONOMIES
################################################################################################

# Fichier Excel
excel_taxonomies = load_workbook(args.excel_taxonomies)
excel_taxonomies_sheets = excel_taxonomies.sheetnames

# Dictionnaire associant l'identifiant Euterpe d'un objet à son UUID Directus
id_uuid = {}

# Création d'une collection Directus par taxonomie
print("\n*** TAXONOMIES ***\n")

for sheet in excel_taxonomies_sheets:
    if sheet == "spécialité":
        # print("SPECIALITES")
        add_id_uuid(sheet)
        # send_taxonomy(sheet, "specialites")
        # print("\n" * 2)
    if sheet == "Période":
        # print("PERIODES")
        add_id_uuid(sheet)
        # send_taxonomy(sheet, "periodes")
        # print("\n" * 2)
    if sheet == "École":
        # print("ECOLES")
        add_id_uuid(sheet)
        # send_taxonomy(sheet, "ecoles")
        # print("\n" * 2)
    if sheet == "Domaine":
        # print("DOMAINES")
        add_id_uuid(sheet)
        # send_taxonomy(sheet, "domaines")
        # print("\n" * 2)
    if sheet == "Lieu de conservation":
        #print("LIEU DE CONSERVATION")
        add_id_uuid(sheet)
        #send_taxonomy(sheet, "lieux_de_conservation")
        #print("\n" * 2)
    if sheet == "Thème":
        #print("THEMES")
        add_id_uuid(sheet)
        #send_tree_taxonomy(sheet, "themes", themes=True)
        #print("\n" * 2)
    if sheet == "Instrument de musique":
        print("INSTRUMENTS DE MUSIQUE")
        add_id_uuid(sheet)
        send_tree_taxonomy(sheet, "instruments_de_musique", themes=False)
        #print("\n" * 2)
    if sheet == "Chant":
        # print("CHANTS")
        add_id_uuid(sheet)
        # send_taxonomy(sheet, "chants")
        # print("\n" * 2)
    if sheet == "Support":
        # print("SUPPORTS")
        add_id_uuid(sheet)
        # send_taxonomy(sheet, "supports")
        # print("\n" * 2)
    if sheet == "Type oeuvre":
        # print("TYPES D'OEUVRES")
        add_id_uuid(sheet)
        # send_taxonomy(sheet, "types_oeuvres")
        # print("\n" * 2)
    if sheet == "Rôles":
        # print("ROLES")
        add_id_uuid(sheet)
        # send_taxonomy(sheet, "roles")
        # print("\n" * 2)
    if sheet == "Notation musicale":
        # print("NOTATION MUSICALE")
        add_id_uuid(sheet)
        # send_taxonomy(sheet, "notations_musicales")
        # print("\n" * 2)

################################################################################################
## DATA
################################################################################################

# Fichier Excel
excel_data = load_workbook(args.excel_data)
excel_sheets = excel_data.sheetnames


# 1. "AUTEURS OEUVRES"
#----------------------------------------------------------

donnees_a_envoyer = []

rows = get_xlsx_sheet_rows_as_dicts(excel_data["1_auteurs"])

# Suppression des items de la collection Directus
# print("\n*** AUTEURS OEUVRES ***\n")
# delete("auteurs_oeuvres_periodes")
# delete("auteurs_oeuvres_specialites")
# delete("auteurs_oeuvres_ecoles")
# delete("auteurs_oeuvres")

for row in rows:

    periodes = []
    specialites = []
    ecoles = []

    # Ajout du lien entre l'identifiant Euterpe de chaque item et son UUID Directus dans un dictionnaire
    id_uuid[str(row["id"])] = row["uuid"]

    # Recherche de l'UUID Directus d'un objet à partir de son identifiant Euterpe
    if row["siècle"] != None:
        get_uuid_list("siècle", periodes)

    if row["spécialité"] != None:
        get_uuid_list("spécialité", specialites)

    if row["école"] != None:
        get_uuid_list("école", ecoles)

    # Body de la requête
    dict = {
        "id": row["uuid"],
        "nom": row["nom"],
        "alias": row["alias"],
        "lieu_de_deces": row["lieu de décès"],
        "periodes": [{
            "periode_id": periode,
            "auteur_oeuvres_id": row["uuid"],
            "collection": "periode"
        } for periode in periodes],
        "specialites": [{
            "specialite_id": specialite,
            "auteur_oeuvres_id": row["uuid"],
            "collection": "specialite"
        } for specialite in specialites],
        "ecoles": [{
            "ecole_id": ecole,
            "auteur_oeuvres_id": row["uuid"],
            "collection": "ecole"
        } for ecole in ecoles],
        "date_de_deces": row["date de décès"],
        "lieu_de_naissance": row["lieu de naissance"],
        "date_de_naissance": row["date de naissance"],
        "commentaire": row["commentaire"],
        "lieu_dactivite": row["lieu d'activité"],
        "date_dactivite": row["date d'activité"]
    }

    donnees_a_envoyer.append(dict)

# Envoi des items dans la collection Directus
# send_data("auteurs_oeuvres", 100, 3300, 3385)


# 2. "OEUVRES LYRIQUES"
#------------------------

donnees_a_envoyer = []

rows = get_xlsx_sheet_rows_as_dicts(excel_data["5_oeuvres_lyriques"])

# Suppression des items de la collection Directus
# print("\n*** OEUVRES LYRIQUES ***\n")
# delete("oeuvres_lyriques_librettistes")
# delete("oeuvres_lyriques_compositeurs")
# delete("oeuvres_lyriques")

for row in rows:

    librettistes = []
    compositeurs = []
    types_oeuvres = []

    # Ajout du lien entre l'identifiant Euterpe de chaque item et son UUID Directus dans un dictionnaire
    id_uuid[str(row["id"])] = row["uuid"]

    # Recherche de l'UUID Directus d'un objet à partir de son identifiant Euterpe
    if row["librettiste"] != None:
        get_uuid_list("librettiste", librettistes)

    if row["compositeur"] != None:
        get_uuid_list("compositeur", compositeurs)

    if row["type_oeuvre"] != None:
        try:
            type_oeuvre_uuid = id_uuid[str(row["type_oeuvre"])]
        except:
            print("type_oeuvre :", row["type_oeuvre"], "not found")

    # Body de la requête
    dict = {
        "id": row["uuid"],
        "titre": row["titre"],
        "librettistes": [{
            "auteur_oeuvres_id": librettiste,
            "oeuvre_lyrique_id": row["uuid"],
            "collection": "auteurs_oeuvres"
        } for librettiste in librettistes],
        "compositeurs": [{
            "auteur_oeuvres_id": compositeur,
            "oeuvre_lyrique_id": row["uuid"],
            "collection": "auteurs_oeuvres"
        } for compositeur in compositeurs],
        "date_oeuvre": row["date_oeuvre"],
        "type": type_oeuvre_uuid,
        "commentaire": row["commentaire"]
    }

    donnees_a_envoyer.append(dict)

# Envoi des items dans la collection Directus
# send_data("oeuvres_lyriques", 100, 100, 123)


# 3. "AUTEURS BIBLIOGRAPHIE"
#---------------------------

donnees_a_envoyer = []

rows = get_xlsx_sheet_rows_as_dicts(excel_data["6_auteurs_bibli_id"])

# Suppression des items de la collection Directus
# print("\n*** AUTEURS BIBLIOGRAPHIE ***\n")
# delete("auteurs_bibliographie")

for row in rows:

    # Ajout du lien entre l'identifiant Euterpe de chaque item et son UUID Directus dans un dictionnaire
    id_uuid[str(row["id"])] = row["uuid"]

    # Body de la requête
    dict = {
        "id": row["uuid"],
        "nom": row["nom"],
        "prenom": row["prénom"]
    }


    donnees_a_envoyer.append(dict)

# Envoi des items dans la collection Directus
# send_data("auteurs_bibliographie", 100, 400, 436)


# 4. "BIBLIOGRAPHIE"
#--------------------

donnees_a_envoyer = []

rows = get_xlsx_sheet_rows_as_dicts(excel_data["3_euterpe_biblio"])

# Suppression des items de la collection Directus
# print("\n*** BIBLIOGRAPHIE ***\n")
# delete("bibliographie_auteurs_bibliographie")
# delete("bibliographie")

for row in rows:

    auteurs = []

    # Ajout du lien entre l'identifiant Euterpe de chaque item et son UUID Directus dans un dictionnaire
    id_uuid[str(row["id"])] = row["uuid"]

    # Recherche de l'UUID Directus d'un objet à partir de son identifiant Euterpe
    if row["auteur_id"] != None:
        get_uuid_list("auteur_id", auteurs)

    # Body de la requête
    dict = {
        "id": row["uuid"],
        "auteurs": [{
            "auteur_bibliographie_id": auteur,
            "bibliographie_id": row["uuid"],
            "collection": "auteurs_bibliographie"
        } for auteur in auteurs],
        "titre": row["titre"],
        "revue_colloque_collection": row["revue_colloque_collection"],
        "parution": row["parution"],
        "editeur": row["editeur"],
        "nbre_n_de_pages": row["nbre_n_de_page"],
        "n_revue_collection": row["n_revue_collection"],
        "lieu_de_publication": row["lieu_d_activite"],
        "commentaire": row["commentaire"]
    }

    donnees_a_envoyer.append(dict)

# Envoi des items dans la collection Directus
# send_data("bibliographie", 100, 700, 721)


# 5. "OEUVRES"
#--------------------

print("\n*** OEUVRES ***\n")

donnees_a_envoyer = []

rows = get_xlsx_sheet_rows_as_dicts(excel_data["4_euterpe_images"])

# Récupération des images de la librairie Directus ("/files")
print("Récupération des images de la librairie Directus\n")
r = requests.get(secret["url"] + f'/files?limit=-1&access_token=' + access_token)
print(r)

# Lien entre l'UUID Directus de l'image et son identifiant Euterpe
images_uuid = {}
for item in r.json()["data"]:
    images_uuid[item["title"]] = item["id"]

# Suppression des items de la collection Directus et items des tables de jointure
print("Suppression des items de la collection Directus et items des tables de jointure :\n")
delete("oeuvres_a_la_maniere_de")
delete("oeuvres_anciennes_attributions")
delete("oeuvres_artistes")
delete("oeuvres_ateliers")
delete("oeuvres_attributions")
delete("oeuvres_copie_dapres")
delete("oeuvres_dapres")
delete("oeuvres_ecoles")
delete("oeuvres_editeurs")
delete("oeuvres_graveurs")
delete("oeuvres_inventeurs")
delete("oeuvres_chants")
delete("oeuvres_domaines")
delete("oeuvres_ecoles")
delete("oeuvres_instruments_de_musique")
delete("oeuvres_lieux_de_conservation")
delete("oeuvres_notations_musicales")
delete("oeuvres_voir_aussi")
delete("oeuvres_oeuvres_representees")
delete("oeuvres_themes")
delete("oeuvres")

# Ajout du lien entre l'identifiant Euterpe de chaque item et son UUID Directus dans un dictionnaire
for row in rows:
    id_uuid[str(row["id"])] = row["uuid"]

# Création d'un dictionnaire par oeuvre (ligne de feuille Excel)
for row in rows:
    # UUIDs des objets indexés lorsqu'il y en a plusieurs par ligne
    editeurs = []
    domaines = []
    instruments = []
    inventeurs = []
    themes = []
    lieux_conservation = []
    notations_musicales = []
    graveurs = []
    artistes = []
    chants = []
    attributions = []
    ecoles = []
    anciennes_attributions = []
    copies = []
    dapres_list = []
    manieres = []
    ateliers = []

    # Recherche de l'UUID Directus d'un objet à partir de son identifiant Euterpe
    if row["éditeur"] != None:
        get_uuid_list("éditeur", editeurs)

    if row["domaine"] != None:
        get_uuid_list("domaine", domaines)

    if row["instrument de musique"] != None:
        get_uuid_list("instrument de musique", instruments)

    if row["inventeur"] != None:
        get_uuid_list("inventeur", inventeurs)

    if row["thème"] != None:
        get_uuid_list("thème", themes)

    if row["lieu de conservation"] != None:
        get_uuid_list("lieu de conservation", lieux_conservation)

    if row["musique écrite"] != None:
        get_uuid_list("musique écrite", notations_musicales)

    if row["graveur"] != None:
        get_uuid_list("graveur", graveurs)

    if row["artiste"] != None:
        get_uuid_list("artiste", artistes)

    if row["chant"] != None:
        get_uuid_list("chant", chants)

    if row["attribution"] != None:
        get_uuid_list("attribution", attributions)

    if row["école"] != None:
        get_uuid_list("école", ecoles)

    if row["ancienne attribution"] != None:
        get_uuid_list("ancienne attribution", anciennes_attributions)

    if row["copie d'après"] != None:
        get_uuid_list("copie d'après", copies)

    if row["d'après"] != None:
        get_uuid_list("d'après", dapres_list)

    if row["manière de"] != None:
        get_uuid_list("manière de", manieres)

    if row["atelier"] != None:
        get_uuid_list("atelier", ateliers)

    # Création d'une date ISO
    try:
        date = re.search("[0-9]{4}", str(row["date œuvre"]))
        date = date.group()[:4]
    except:
        date = None

    # Body de la requête
    dict = {
        "id": row["uuid"],
        "titre": row["titre"],
        "titre_alternatif": row["titre alternatif"],
        "editeur": [{
            "auteur_oeuvre_id": editeur,
            "oeuvre_id": row["uuid"]
        } for editeur in editeurs],
        "reference_iremus": row["référence iremus"].replace("🍄", ",") if row["référence iremus"] != None else row["référence iremus"],
        "domaines": [{
            "domaine_id": domaine,
            "oeuvre_id": row["uuid"]
        } for domaine in domaines],
        "num_inventaire": row["n° inventaire"],
        "cote": row["cote"],
        "inscription": row["inscription"],
        "technique": row["technique"],
        "oeuvre_en_rapport": row["œuvre en rapport"],
        "themes": [{
            "theme_id": theme,
            "oeuvre_id": row["uuid"]
        } for theme in themes],
        "inventeur": [{
            "auteur_oeuvre_id": inventeur,
            "oeuvre_id": row["uuid"]
        } for inventeur in inventeurs],
        "lieux_de_conservation": [{
            "lieu_de_conservation_id": lieu_conservation,
            "oeuvre_id": row["uuid"]
        } for lieu_conservation in lieux_conservation],
        "date": row["date œuvre"],
        "date_iso": date,
        "precision_oeuvre": row["précision œuvre"],
        "precision_instrument": row["précision instrument"],
        "notation_musicale": [{
            "notation_musicale_id": notation_musicale,
            "oeuvre_id": row["uuid"]
        } for notation_musicale in notations_musicales],
        "graveur": [{
            "auteur_oeuvre_id": graveur,
            "oeuvre_id": row["uuid"]
        } for graveur in graveurs],
        "commentaire": row["commentaire"],
        "bibliographie": row["bibliographie"],
        "reference_agence": row["référence agence"],
        "hauteur": row["hauteur"],
        "largeur": row["largeur"],
        "diametre": row["diamètre"],
        "chants": [{
            "chant_id": chant,
            "oeuvre_id": row["uuid"]
        } for chant in chants],
        "artistes": [{
            "auteur_oeuvre_id": artiste,
            "oeuvre_id": row["uuid"]
        } for artiste in artistes],
        "ecoles": [{
                "auteur_oeuvre_id": ecole,
                "oeuvre_id": row["uuid"]
            } for ecole in ecoles],
        "attributions": [{
                    "auteur_oeuvre_id": attribution,
                    "oeuvre_id": row["uuid"]
                } for attribution in attributions],
        "precision_musique": row["précision musique"],
        "anciennes_attributions": [{
            "auteur_oeuvre_id": ancienne_attribution,
            "oeuvre_id": row["uuid"]
        } for ancienne_attribution in anciennes_attributions],
        "source_litteraire": row["source littéraire"],
        "copie_dapres": [{
            "auteur_oeuvre_id": copie,
            "oeuvre_id": row["uuid"]
        } for copie in copies],
        "dapres": [{
            "auteur_oeuvre_id": dapres,
            "oeuvre_id": row["uuid"]
        } for dapres in dapres_list],
        "a_la_maniere_de": [{
            "auteur_oeuvre_id": maniere,
            "oeuvre_id": row["uuid"]
        } for maniere in manieres],
        "instruments_de_musique": [{
            "instrument_de_musique_id": instrument,
            "oeuvre_id": row["uuid"]
        } for instrument in instruments],
        "ateliers": [{
            "atelier_id": atelier,
            "oeuvre_id": row["uuid"]
        } for atelier in ateliers]
    }

    # URL
    if row["url"] != None:
        if row["titre de l'url"] != None:
            dict["url"] = "<a href=" + row["url"] + ">" + row["titre de l'url"] + "</a>"
        else:
            dict["url"] = "<a href=" + row["url"] + ">Lien</a>"

    # Images
    if row["image"] != None:
        images_unsplit = str(row["image"]).split("🍄")
        images = [image.strip() for image in images_unsplit]
        try :
            dict["images"] = [{
                "image_id": images_uuid[image],
                "oeuvre_id": row["uuid"]
            } for image in images]
        except:
            print("image:", row["image"], ": image non trouvée")

        # Sélection d'une image pour la miniature
        miniature = images[0]
    try :
        dict["miniature"] = images_uuid[miniature]
    except:
        print("image:", miniature, ": image non trouvée")

    # "contient/contenu dans" à écrire à la main? (un seul enregistrement)

#     donnees_a_envoyer.append(dict)
#
# print("Ecriture du fichier JSON\n")
# with open(args.oeuvres_a_envoyer, "w") as f:
#     json.dump(donnees_a_envoyer, f, ensure_ascii=False)

# Envoi des items dans la collection Directus par paquets de 200
print("Envoi de 300 items")
with open(args.oeuvres_a_envoyer, "r") as json_oeuvres_a_envoyer:
    oeuvres_a_envoyer = json.load(json_oeuvres_a_envoyer)
    paquet = oeuvres_a_envoyer[0:300]
    for i in range(0, len(paquet), 1):
        try:
            r = requests.post(secret["url"] + f'/items/oeuvres?limit=-1&access_token=' + access_token, json=paquet[i])
            print(i)
        except Exception as e:
            print("titre de l'oeuvre :", paquet[i]["titre"])
            print(r.json(), "\n")

# Ajout des informations d'une collection faisant référence à elle-même (PATCH)
# print("\nAjout des 'voir aussi' et 'oeuvres représentées' (requête PATCH):\n")
#
# infos_a_patcher = []
#
# for row in rows:
#     if row["voir aussi"] != None or row["œuvre représentée"] != None:
#         voir_aussi = []
#         items_a_envoyer = []
#
#         id_oeuvre = id_uuid[str(row["id"])]
#
#         # Voir aussi
#         if row["voir aussi"] != None:
#             get_uuid_list("voir aussi", voir_aussi)
#             item = {}
#             item["id"] = id_oeuvre
#             item["voir_aussi"] = [{
#                 "voir_aussi_id": v,
#                 "oeuvre_id": row["uuid"]
#             } for v in voir_aussi]
#
#             items_a_envoyer.append(item)
#
#         # Oeuvres représentées
#         if row["œuvre représentée"] != None:
#             oeuvres_representees_unsplit = row["œuvre représentée"].split("🍄")
#             oeuvres_representees = [oeuvre_representee.strip() for oeuvre_representee in oeuvres_representees_unsplit]
#             images_representees = []
#             oeuvres_lyriques_representees = []
#
#             for oeuvre_representee in oeuvres_representees:
#                 try:
#                     images_representees.append(images_uuid[str(oeuvre_representee)])
#                 except:
#                     try:
#                         oeuvres_lyriques_representees.append(id_uuid[str(oeuvre_representee)])
#                     except:
#                         print(row["titre"], ":")
#                         print(oeuvre_representee, "non trouvée")
#
#             if len(images_representees) >= 1:
#                 item = {}
#                 item["id"] = id_oeuvre
#                 item["oeuvres_representees"] = [{
#                     "item": image,
#                     "oeuvre_id": row["uuid"],
#                     "collection": "directus_files"} for image in images_representees]
#                 items_a_envoyer.append(item)
#
#             if len(oeuvres_lyriques_representees) >= 1:
#                 item = {}
#                 item["id"] = id_oeuvre
#                 item["oeuvres_representees"] = [{
#                     "item": oeuvre_lyrique,
#                     "oeuvre_id": row["uuid"],
#                     "collection": "oeuvres_lyriques"} for oeuvre_lyrique in oeuvres_lyriques_representees]
#                 items_a_envoyer.append(item)

        # for item in items_a_envoyer:
        #     try:
        #         pprint(item)
        #         r = requests.patch(secret["url"] + '/items/oeuvres/' + id_oeuvre + '?access_token=' + access_token, json=item)
        #         print(r, "\n")
        #     except Exception as e:
        #         print(e)
        #         pprint(r.json(), "\n")


