import os, sys
import requests
import yaml
import argparse
from openpyxl import load_workbook
from pprint import pprint
import time
import re
from pprint import pprint
from geopy.geocoders import Nominatim

# Helpers
sys.path.append(os.path.abspath(os.path.join('python_packages/helpers_excel', '')))
from helpers_excel import *

# Arguments
parser = argparse.ArgumentParser()
parser.add_argument("--excel_data")
args = parser.parse_args()

############################################################################################
# EXTRACTING THE INFORMATION FROM THE EXCEL SHEETS
############################################################################################

# Reading the Excel file
excel_data = load_workbook(args.excel_data)
excel_data_sheets = excel_data.sheetnames

# Data
data = []

# Adding both sheets of the excel file in the "data" list of dictionaries.
print("\nEXTRACTING THE INFORMATION FROM THE EXCEL SHEETS\n")
for s in excel_data_sheets:
    rows = get_xlsx_sheet_rows_as_dicts(excel_data[s])
    for row in rows:
        dict = {
            "arrondissement": row["Arrondissement"],
            "nom_de_la_salle": row["Nom de la salle"],
            "raison_sociale_de_l_etablissement" : row["Raison sociale de l'établissement"],
            "type_de_lieu" : row["Type de lieu"],
            "adresse" : row["Adresse"],
            "date_de_construction" : row["Date de construction"],
            "date_de_premiere_activite_de_concert" : row["Date de première activité de concert"],
            "date_de_derniere_inauguration" : row["Date de dernière inauguration"],
            "exploitant" : row["Exploitant"],
            "statut_juridique" : row["Statut juridique"],
            "proprietaire" : row["Propriétaire"],
            "reseau" : row["Réseau"],
            "jazz_et_blues" : row["Jazz et blues"],
            "pop_et_rock" : row["Pop et Rock"],
            "rnb_rap_soul" : row["RnB, Rap, Soul"],
            "musique_electronique" : row["Musique électronique"],
            "musique_du_monde" : row["Musique du monde"],
            "chanson_et_variete" : row["Chanson et variété"],
            "gospel" : row["Gospel"],
            "musique_chorale" : row["Musique chorale"],
            "theatre_musical" : row["Théâtre musical"],
            "musique_symphonique" : row["Musique symphonique"],
            "musique_de_chambre" : row["Musique de chambre"],
            "opera" : row["Opéra"],
            "recital" : row["Récital"],
            "cine_concerts" : row["Ciné-concerts"],
            "autres" : row["Autres (préciser quoi dans la colonne)"],
            "jauge assise" : row["Jauge assise"],
            "jauge_debout" : row["Jauge debout"],
            "salle_privatisable" : row["La salle est-elle privatisable ?"],
            "modes_de_production" : row["Modes de production"],
            "actions_culturelles" : row["Actions culturelles (ateliers, formations, conférences, cours, etc.)"],
            "espace_de_sociabilite_en_libre_acces" : row["Espace de sociabilité en libre accès (hall, cours, terrasse...)"],
            "espace_de_restauration_payant" : row["Espace de restauration (payant)"],
            "presence_d_un_centre_de_ressources" : row["Présence d'un centre de ressources (bibliothèque, médiathèque, etc.)"],
            "presence_d_espaces_privatisables" : row["Présence d'espaces privatisables"],
            "acces_pmr": row["Accès PMR"],
            "subventionne": row["Subventionné"],
            "notice": row["Notice"],
            "dimensions_plateau" : row["Dimensions plateau"],
            "dimensions_salle" : row["Dimensions salle"],
            "commentaires": row["Commentaires"],
            "contact" : row["Contact"]
        }

        # Parisian places
        if "Département" not in row:
            dict["departement"] = 75
            dict["commune"] = "Paris"
        else:
            dict["departement"] = row["Département"]
            dict["commune"] = row["Commune"]

        # Website
        if row["site internet"] != None:
            dict["site_internet"] = "<a href=" + row["site internet"] + ">Lien</a>"

        # Standardised date
        def create_standard_date(date_in_excel_sheet, name_of_standardised_field_in_directus):
            if date_in_excel_sheet != None:
                date_in_excel_sheet = str(date_in_excel_sheet).replace(" ", "")
                # Finding four digit dates
                four_digit_date = re.search("[0-9]{4}", date_in_excel_sheet)
                if four_digit_date != None:
                    # Date interval
                    if "-" in date_in_excel_sheet:
                        interval = date_in_excel_sheet.split("-")
                        if len(interval[1]) < 3:
                            interval[1] = interval[0][0:2] + interval[1]
                        interval = interval[0] + "," + interval[1]
                        dict[name_of_standardised_field_in_directus] = interval
                    else:
                        dict[name_of_standardised_field_in_directus] = four_digit_date.group()
                else:
                    # Finding three digit dates
                    three_digit_date = re.search("[0-9]{3}", date_in_excel_sheet)
                    if three_digit_date != None:
                        # Date interval
                        if "-" in date_in_excel_sheet:
                            interval = date_in_excel_sheet.split("-")
                            if len(interval[1]) < 3:
                                interval[1] = interval[0][0:2] + interval[1]
                            interval = interval[0] + "," + interval[1]
                            dict[name_of_standardised_field_in_directus] = interval
                        else:
                            dict[name_of_standardised_field_in_directus] = three_digit_date.group()
                    else:
                        pass

        create_standard_date(row["Date de construction"], "date_de_construction_iso")
        create_standard_date(row["Date de première activité de concert"], "date_de_premiere_activite_de_concert_iso")

        # Geolocation
        # Creating geolocator object
        geolocator = Nominatim(user_agent="Iremus")

        if row["Adresse"] != None and row["Adresse"] != "?" \
            and row["Adresse"] != "NC" and row["Adresse"] != "Non spécifié":
            try:
                location = geolocator.geocode(row["Adresse"])
                latitude = str(location.latitude)
                longitude = str(location.longitude)
                dict["coordonnees_geographiques"] = latitude + ", " + longitude
                print("Concert place address geolocated successfully")
            except:
                print(row["Adresse"], ": Geopy couldn't find any coordinates for this address")
                pass

        data.append(dict)


############################################################################################
# INSERTING THE INFORMATION INTO THE DIRECTUS COLLECTION
############################################################################################

# YAML Secret
file = open(os.path.join(sys.path[0], "secret.yaml"))
secret = yaml.full_load(file)
r = requests.post(secret["url"] + "/auth/login",
                  json={"email": secret["email"], "password": secret["password"]})
access_token = r.json()['data']['access_token']
refresh_token = r.json()['data']['refresh_token']
file.close()

# Sending the data in paquets of 10
print("\n SENDING THE INFORMATION TO DIRECTUS COLLECTION\n")
print(len(data), "items to send")
for i in range(0, len(data), 50):
    print(i)
    try:
        r = requests.post(secret["url"] + f'/items/lieux_de_concert?limit=-1&access_token=' + access_token, json=data[i:i+50])
        r.raise_for_status()
    except Exception as e:
        print(e)
        print(data[i:i+50])
        print(r.json())

# # Sending the remaining data
for i in range(950, 957):
    print(i)
    try:
        r = requests.post(secret["url"] + f'/items/lieux_de_concert?limit=-1&access_token=' + access_token, json=data[i])
        r.raise_for_status()
    except Exception as e:
        print(e)
        pprint(data[i])
        print(r.json())